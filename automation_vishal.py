
from openpyxl import load_workbook
from datetime import date, datetime
import os
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.styles import Font
from openpyxl.styles import Alignment


# ================= FILE PATHS =================
SOURCE_FILE_PATH = ''
TARGET_FILE_PATH =''
inventory_file_path =''




SOURCE_FILE = ''
TARGET_FILE = ''
INVENTORY_FILE = ''
# ================= SHEETS =================
source_sheet = ''
target_sheet = ''
inventory_sheet = ''




thin = Side(style="thin", color="000000")
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
abadi_font = Font(
    name="Abadi",
    size=11
)

carrier_color = PatternFill(
    fill_type="solid",
    fgColor="FFC0E6F5"
)

backbone_color = PatternFill(
    fill_type="solid",
    fgColor="FFF2DCDB"
)

center_alignment = Alignment(horizontal="center", vertical="center")

# ================= GLOBAL =================

# ================= HELPERS =================
inventory_mapping_dict = {}
date_router={}

def initialize_path(source_file_name, target_file_name, inventory_file_name):
    global SOURCE_FILE_PATH, TARGET_FILE_PATH, INVENTORY_FILE_PATH
    SOURCE_FILE_PATH = os.path.join(os.getcwd(), "maintainance", source_file_name)
    TARGET_FILE_PATH = os.path.join(os.getcwd(), "maintainance", target_file_name)
    INVENTORY_FILE_PATH = os.path.join(os.getcwd(), "maintainance", inventory_file_name)


    global SOURCE_FILE, TARGET_FILE, INVENTORY_FILE, source_sheet, target_sheet, inventory_sheet
    SOURCE_FILE = load_workbook(SOURCE_FILE_PATH, data_only=True)
    TARGET_FILE = load_workbook(TARGET_FILE_PATH)
    INVENTORY_FILE = load_workbook(INVENTORY_FILE_PATH, data_only=True)
    # ================= SHEETS =================
    source_sheet = SOURCE_FILE["Scheduled Maintenance"]
    target_sheet = TARGET_FILE["OBS Maintenance Tracker"]
    inventory_sheet = INVENTORY_FILE["Haleon WAN Inventory"]


def inventory_mapping():
    # Implementation for inventory mapping

    for row in inventory_sheet.iter_rows(min_row=2, values_only=True):
        inventory_mapping_dict[row[4]] = {
            "TOPO":row[6],
            "ROLE":row[7],
        }  # Assuming column 5 is device and column 1 is site

def same_day_maintenance():
    for formatted_date in date_router.keys():
        for router_name in date_router[formatted_date].keys():
            rows = date_router[formatted_date][router_name]
            if router_name in inventory_mapping_dict and inventory_mapping_dict[router_name]["TOPO"]:
                topo=inventory_mapping_dict[router_name]["TOPO"].split(".")
                if len(topo) > 1:
                    secondary_device = ""
                    if(topo[0]==router_name):
                        secondary_device = topo[1]
                    else:
                        secondary_device = topo[0]
                    if secondary_device in date_router[formatted_date].keys():
                        for row in rows:
                            target_sheet.cell(row=row, column=14).value = "YES"
                            target_sheet.cell(row=row, column=14).fill = PatternFill(fill_type="solid", fgColor="FFFF0000")  # Red color for "YES"
                            target_sheet.cell(row=row, column=15).value = "Either a router has 2 or more maintenace on same day OR primary and secondary routers are under maintenace on same day, please check the failover device and plan accordingly"

            if len(rows) > 1:
                for row in rows:
                        target_sheet.cell(row=row, column=14).value = "YES"
                        target_sheet.cell(row=row, column=14).fill = PatternFill(fill_type="solid", fgColor="FFFF0000")  # Red color for "YES"
                        target_sheet.cell(row=row, column=15).value = "Either a router has 2 or more maintenace on same day OR primary and secondary routers are under maintenace on same day, please check the failover device and plan accordingly"

                print(f"Router {router_name} not found in inventory mapping.")



def copy():
    inventory_mapping()
    max_len=source_sheet.max_row
    count=1 
    for c_row in range(12, max_len + 1):
        target_row=c_row-1
        target_sheet.cell(row=target_row, column=3, value=source_sheet.cell(row=c_row, column=1).value )
         # A -> C
        target_sheet.cell(row=target_row, column=3).value = source_sheet.cell(row=c_row, column=1).value

        # B -> D
        target_sheet.cell(row=target_row, column=4).value = source_sheet.cell(row=c_row, column=2).value

        # C -> E
        target_sheet.cell(row=target_row, column=5).value = source_sheet.cell(row=c_row, column=3).value

        # P -> F
        p_cell = source_sheet.cell(row=c_row, column=1)

        cell_color = p_cell.fill.fgColor.value
        #print(p_cell.fill.fgColor.value)
        
        #Serial no 
        target_sheet.cell(row=target_row, column=2).value = count
        count += 1

        #check for backbone and carrier maintenance based on the color of the cell in column P
        if  cell_color == "FFF2DCDB":
            target_sheet.cell(row=target_row, column=6).value = "Backbone Maintenance"
        else:
            target_sheet.cell(row=target_row, column=6).value = "Carrier Maintenance"

        # D -> G
        target_sheet.cell(row=target_row, column=7).value = source_sheet.cell(row=c_row, column=4).value

        # E -> H
        target_sheet.cell(row=target_row, column=8).value = source_sheet.cell(row=c_row, column=5).value

        # F -> I
        target_sheet.cell(row=target_row, column=9).value = source_sheet.cell(row=c_row, column=6).value

        # G -> J
        target_sheet.cell(row=target_row, column=10).value = source_sheet.cell(row=c_row, column=7).value


        # J -> K
        target_sheet.cell(row=target_row, column=11).value = source_sheet.cell(row=c_row, column=10).value

        # K -> L
        target_sheet.cell(row=target_row, column=12).value = source_sheet.cell(row=c_row, column=11).value

        # L -> M
        target_sheet.cell(row=target_row, column=13).value = source_sheet.cell(row=c_row, column=12).value

        target_sheet.cell(row=target_row, column=16).value = "YES"
        target_sheet.cell(row=target_row, column=17).value = "YES"
        target_sheet.cell(row=target_row, column=18).value = "YES"


        target_sheet.cell(row=target_row, column=21).value = source_sheet.cell(row=c_row, column=21).value

       # dictionaty making phase for the date and router name
        router_name=source_sheet.cell(row=c_row, column=2).value
        date=source_sheet.cell(row=c_row, column=7).value
        formatted_date = datetime.strptime(date, "%d/%b/%Y %I:%M:%S %p").strftime("%d/%m/%Y")

        if formatted_date not in date_router:
            date_router[formatted_date] = {}
        if router_name not in date_router[formatted_date]:
            date_router[formatted_date][router_name] = []
        date_router[formatted_date][router_name].append(target_row)
        

        # upating top left box ( meta info about the sheet)
         # Source E3 -> Target C3
        target_sheet["C3"] = source_sheet["E3"].value

        # Source E4 -> Target C4
        target_sheet["C4"] = source_sheet["E4"].value

        # Source E5 -> Target C6
        target_sheet["C6"] = source_sheet["E5"].value

        # Source E6 -> Target C7
        target_sheet["C7"] = source_sheet["E6"].value

        # Source E7 -> Target C8
        target_sheet["C8"] = source_sheet["E7"].value

                # Apply formatting to the target sheet
        for col in range(1, 24):
            cell = target_sheet.cell(row=target_row, column=col)
            cell.border = thin_border
            target_sheet.row_dimensions[target_row].height = 30
            cell.font = abadi_font
            cell.fill = carrier_color if target_sheet.cell(row=target_row, column=6).value == "Carrier Maintenance" else backbone_color
            cell.alignment = center_alignment


        if source_sheet.cell(row=c_row, column=2).value in inventory_mapping_dict.keys() and inventory_mapping_dict[source_sheet.cell(row=c_row, column=2).value]["TOPO"]!= None:
            role = inventory_mapping_dict[source_sheet.cell(row=c_row, column=2).value]["ROLE"]
            topo = inventory_mapping_dict[source_sheet.cell(row=c_row, column=2).value]["TOPO"].split(".")
            if str.upper(role) == "UNDERLAY":
                if len(topo) > 1:
                    secondary_device = ""
                    print(topo[0]+ "    "+ topo[1] + "    " + source_sheet.cell(row=c_row, column=2).value)
                    if(topo[0]==source_sheet.cell(row=c_row, column=2).value):
                        secondary_device = topo[1]
                    else:
                        secondary_device = topo[0]
                    target_sheet.cell(row=target_row, column=15).value = f"Traffic from Underlay {source_sheet.cell(row=c_row, column=2).value} will be failover to {secondary_device}."
                    target_sheet.cell(row=target_row, column=14).value = "NO"
                    target_sheet.cell(row=target_row, column=14).fill = PatternFill(fill_type="solid", fgColor="FF00FF00")  # Green color for "NO"
                else:
                    target_sheet.cell(row=target_row, column=15).value = "No secondary underlay defined in inventory. Please update the inventory or connect with mayank or gurmeet"
                    target_sheet.cell(row=target_row, column=14).value = "UNKNOWN"
                    target_sheet.cell(row=target_row, column=14).fill = PatternFill(fill_type="solid", fgColor="FF0000FF")  # Blue color for "UNKNOWN"
            elif len(topo) > 1:
                secondary_device = ""
                print(topo[0]+ "    "+ topo[1] + "    " + source_sheet.cell(row=c_row, column=2).value)
                if(topo[0]==source_sheet.cell(row=c_row, column=2).value):
                    secondary_device = topo[1]
                else:
                    secondary_device = topo[0]
                target_sheet.cell(row=target_row, column=15).value = f"Traffic from {source_sheet.cell(row=c_row, column=2).value} will be failover to {secondary_device}"
                target_sheet.cell(row=target_row, column=14).value = "NO"
                target_sheet.cell(row=target_row, column=14).fill = PatternFill(fill_type="solid", fgColor="FF00FF00")  # Green color for "NO"

            elif str.upper(role) == 'SINGLE':
                target_sheet.cell(row=target_row, column=15).value = "no secondary router or failover device is available to manage and sustain network traffic"
                target_sheet.cell(row=target_row, column=14).value = "YES"
                target_sheet.cell(row=target_row, column=14).fill = PatternFill(fill_type="solid", fgColor="FFFF0000")  # Red color for "YES"

            else:
                target_sheet.cell(row=target_row, column=15).value = "NO SUCH DEVICE FOUND IN INVENTORY: Please check the device name and update the inventory or connect with mayank or gurmeet"
                target_sheet.cell(row=target_row, column=14).value = "UNKNOWN"
                target_sheet.cell(row=target_row, column=14).fill = PatternFill(fill_type="solid", fgColor="FF0000FF")  # Blue color for "UNKNOWN"
        else:
            print("outer most else loop")
            target_sheet.cell(row=target_row, column=15).value = "NO SUCH DEVICE FOUND IN INVENTORY: Please check the device name and update the inventory or connect with mayank or gurmeet"
            target_sheet.cell(row=target_row, column=14).value = "UNKNOWN"
            target_sheet.cell(row=target_row, column=14).fill = PatternFill(fill_type="solid", fgColor="FF0000FF")  # Blue color for "UNKNOWN"

    same_day_maintenance()

                                                                                       

        


# ================= MAIN =================

def orchestration(source_file_name, target_file_name, inventory_file_name):
    initialize_path(source_file_name, target_file_name, inventory_file_name)
    copy()
    output_file = f"Haleon_Output_{date.today().isoformat()}.xlsx"
    output_path = os.path.join(os.getcwd(), "maintainance", output_file)

    TARGET_FILE.save(output_path)
    print(f"Saved: {output_path}")


