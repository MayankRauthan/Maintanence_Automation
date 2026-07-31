from openpyxl import load_workbook
from datetime import date
import os
import cvmt_extraction as ce, machx_extraction as me
from playwright.sync_api import Playwright, sync_playwright
from openpyxl.styles import PatternFill





# Define file paths
SOURCE_FILE_PATH = os.path.join(os.getcwd(), "maintainance", "Maintenance-Data(125).xlsx")
TARGET_FILE_PATH= os.path.join(os.getcwd(), "maintainance", "Haleon -Network planned maintenance tracker Week 13_25 March_2026.xlsx")
INVENTORY_FILE_PATH = os.path.join(os.getcwd(), "maintainance", "Haleon WAN Inventory.xlsx")

# Load the workbooks
SOURCE_FILE = load_workbook(SOURCE_FILE_PATH, data_only=True)
TARGET_FILE = load_workbook(TARGET_FILE_PATH, data_only=True)
INVENTORY_FILE = load_workbook(INVENTORY_FILE_PATH, data_only=True)

# Load the sheets
# Load target sheets
target_sheet_scheduled_maintenance = TARGET_FILE["Scheduled Maintenance"]
target_sheet_completed_maintenance = TARGET_FILE["Completed Maintenance"]

# Load inventory sheet
inventory_sheet = INVENTORY_FILE["Haleon WAN Inventory"]    

#Load source sheets
source_sheet_orange = SOURCE_FILE["Orange"]
source_sheet_carrier = SOURCE_FILE["Carrier"]

# save the file name with date added ( YYYY-MM-DD ) to avoid overwriting previous outputs
output_file = f"Haleon_Automated_Maintenance_Output_{date.today().isoformat()}.xlsx"

#initialize the mapping dictionary for device to site mapping
dict_device_to_site = {}

orange_mapping_dict = {}

carrier_mapping_dict = {}
CARRIER_FILL = PatternFill(fill_type="solid", fgColor="DCE6F1")   # light blue
ORANGE_FILL = PatternFill(fill_type="solid", fgColor="F2DCDB")  # light pink


#convert column letter to number ( A -> 0, B -> 1, C -> 2, ... ) useful for .cell function in openpyxl which uses column numbers starting from 1
def no(val):
    return ord(val.upper()) - 65

# maps device name to site name using the inventory sheet and stores it in a dictionary for quick lookup 
def inventory_mapping():
    for row in inventory_sheet.iter_rows(min_row=2, values_only=True):
        device_name = row[4].strip().upper()
        site_name = row[0].strip().upper()
        #print ( f"mapping device {device_name} to site {site_name}")
        dict_device_to_site[device_name] = site_name

def orange_mapping():
    global orange_mapping_dict
    orange_mapping_dict = {
    "orange_ref": 1,          # A
    "device_name": 6,        # F
    "site_id": 8,             # H
    "city": 9,                # I
    "country": 10,           # J
    "region": 11,            # K
    "maint_type": 12,        # L
    "sched_gmt": 15,         # O
    "local_time": 17,        # Q
    "local_tz": 18,          # R
    "duration": 19,          # S
    "window": 20,            # T
    "service_impact": 22,    # V
    "orange_status": 23,     # W
    "scope": 24,             # X
}

def carrier_mapping():
   global carrier_mapping_dict
   carrier_mapping_dict = {
    "orange_ref": 1,          # A (row[0])
    "device_name": 3,        # C (row[2])
    "site_id": 8,             # H (row[7])
    "city": 9,                # I (row[8])
    "country": 10,           # J (row[9])
    "region": 11,            # K (row[10])
    "maint_type": 12,        # L (row[11])
    "sched_gmt": 15,         # O (row[14])
    "local_time": 17,        # Q (row[16])
    "local_tz": 18,          # R (row[17])
    "duration": 19,          # S (row[18])
    "window": 20,            # T (row[19])
    "service_impact": 22,    # V (row[21])
    "orange_status": 23,     # W (row[22])
    "carrier_name": 26,     # Z (row[25])
    "carrier_circuit": 27,  # AA (row[26])
    "carrier_ref": 28,      # AB (row[27])
    "carrier_status": 29,   # AC (row[28])
    "reason": 30,            # AD (row[29])
}
   
def extract_status():
    with sync_playwright() as p:
        browser = p.firefox.launch(headless=False)
        context = browser.new_context()
        for i in range (8, source_sheet_orange.max_row + 1):
            cell_value = source_sheet_orange.cell(row=i, column=34).value
            if cell_value:
                print (f"Status found in Orange sheet at row {i}: {cell_value}")
                print(me.get_data(context, cell_value))
            

def fill_row(sheet, row, start_col, end_col, fill):
    for col in range(start_col, end_col + 1):
        sheet.cell(row=row, column=col).fill = fill

def orange_data_extraction():
    target_row = 12
    with sync_playwright() as p:
        browser = p.firefox.launch(headless=False)
        context = browser.new_context()

        
        for i in range(8, source_sheet_orange.max_row + 1):
            duration_value = source_sheet_orange.cell(
                row=i,
                column=orange_mapping_dict["duration"]
            ).value

            # Skip rows where duration is literal text "NONE"
            if duration_value is not None and str(duration_value).strip().upper() == "NONE":
                continue
            
            device_name = source_sheet_orange.cell(row=i, column=orange_mapping_dict["device_name"]).value
            site_id = source_sheet_orange.cell(row=i, column=orange_mapping_dict["site_id"]).value
            scope = source_sheet_orange.cell(row=i, column=orange_mapping_dict["scope"]).value
            city = source_sheet_orange.cell(row=i, column=orange_mapping_dict["city"]).value
            country = source_sheet_orange.cell(row=i, column=orange_mapping_dict["country"]).value
            region = source_sheet_orange.cell(row=i, column=orange_mapping_dict["region"]).value
            sched_gmt = source_sheet_orange.cell(row=i, column=orange_mapping_dict["sched_gmt"]).value
            local_time = source_sheet_orange.cell(row=i, column=orange_mapping_dict["local_time"]).value
            local_tz = source_sheet_orange.cell(row=i, column=orange_mapping_dict["local_tz"]).value
            duration = source_sheet_orange.cell(row=i, column=orange_mapping_dict["duration"]).value
            window = source_sheet_orange.cell(row=i, column=orange_mapping_dict["window"]).value
            service_impact = source_sheet_orange.cell(row=i, column=orange_mapping_dict["service_impact"]).value
            orange_ref = source_sheet_orange.cell(row=i, column=orange_mapping_dict["orange_ref"]).value
            maint_type = source_sheet_orange.cell(row=i, column=orange_mapping_dict["maint_type"]).value
            orange_status = source_sheet_orange.cell(row=i, column=orange_mapping_dict["orange_status"]).value
            orange_link= source_sheet_orange.cell(row=i, column=34).value

            # Site name from inventory, fallback to site_id from orange sheet
            site_name = None
            if device_name is not None:
                site_name = dict_device_to_site.get(str(device_name).strip().upper())

            if not site_name and site_id is not None:
                site_name = site_id
            else:
                site_name=city

            
            # Write to Scheduled Maintenance sheet
            target_sheet_scheduled_maintenance.cell(row=target_row, column=1, value=site_name)        # A Site Name
            target_sheet_scheduled_maintenance.cell(row=target_row, column=2, value=device_name)      # B Device Name
            target_sheet_scheduled_maintenance.cell(row=target_row, column=3, value=scope)            # C Scope of Change
            target_sheet_scheduled_maintenance.cell(row=target_row, column=4, value=city)             # D City
            target_sheet_scheduled_maintenance.cell(row=target_row, column=5, value=country)          # E County/Country
            target_sheet_scheduled_maintenance.cell(row=target_row, column=6, value=region)           # F Region
            target_sheet_scheduled_maintenance.cell(row=target_row, column=7, value=sched_gmt)        # G Scheduled Date GMT
            target_sheet_scheduled_maintenance.cell(row=target_row, column=9, value=local_time)       # I Scheduled Date Local
            target_sheet_scheduled_maintenance.cell(row=target_row, column=10, value=local_tz)        # J Local Time Zone
            target_sheet_scheduled_maintenance.cell(row=target_row, column=11, value=duration)        # K Duration
            target_sheet_scheduled_maintenance.cell(row=target_row, column=12, value=window)          # L Maintenance Window
            target_sheet_scheduled_maintenance.cell(row=target_row, column=13, value=service_impact)  # M Service Impact
            target_sheet_scheduled_maintenance.cell(row=target_row, column=14, value=orange_ref)      # N Orange Change Reference
            target_sheet_scheduled_maintenance.cell(row=target_row, column=18, value=maint_type)      # R Maintenance Type
            target_sheet_scheduled_maintenance.cell(row=target_row, column=20, value=orange_status)   # T Orange Status
            target_sheet_scheduled_maintenance.cell(row=target_row, column=24, value=orange_link)     # U Orange Link
            fill_row(target_sheet_scheduled_maintenance, target_row, 1, 23, ORANGE_FILL)
           # target_sheet_scheduled_maintenance.cell(row=target_row,column=21,value=me.get_data(context, orange_link))
            target_row += 1
            
    return target_row

def carrier_data_extraction(start_row):
    target_row = start_row
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context()

        for i in range(8, source_sheet_carrier.max_row + 1):
            device_name = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["device_name"]).value
            site_id = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["site_id"]).value
            scope = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["reason"]).value
            city = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["city"]).value
            country = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["country"]).value
            region = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["region"]).value
            sched_gmt = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["sched_gmt"]).value
            local_time = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["local_time"]).value
            local_tz = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["local_tz"]).value
            duration = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["duration"]).value
            window = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["window"]).value
            service_impact = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["service_impact"]).value
            orange_ref = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["orange_ref"]).value
            carrier_name = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["carrier_name"]).value
            carrier_circuit = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["carrier_circuit"]).value
            carrier_ref = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["carrier_ref"]).value
            carrier_status = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["carrier_status"]).value
            maint_type = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["maint_type"]).value
            orange_status = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["orange_status"]).value
            carrier_link = source_sheet_carrier.cell(row=i, column=34).value

            # Site name from inventory, fallback to site_id from carrier sheet
            site_name = None
            if device_name is not None:
                site_name = dict_device_to_site.get(str(device_name).strip().upper())

            if not site_name and site_id is not None:
                site_name = site_id
            else:
                site_name=city

            # Write to Scheduled Maintenance sheet
            target_sheet_scheduled_maintenance.cell(row=target_row, column=1, value=site_name)         # A Site Name
            target_sheet_scheduled_maintenance.cell(row=target_row, column=2, value=device_name)       # B Device Name
            target_sheet_scheduled_maintenance.cell(row=target_row, column=3, value=scope)             # C Scope of Change
            target_sheet_scheduled_maintenance.cell(row=target_row, column=4, value=city)              # D City
            target_sheet_scheduled_maintenance.cell(row=target_row, column=5, value=country)           # E County/Country
            target_sheet_scheduled_maintenance.cell(row=target_row, column=6, value=region)            # F Region
            target_sheet_scheduled_maintenance.cell(row=target_row, column=7, value=sched_gmt)         # G Scheduled Date GMT
            target_sheet_scheduled_maintenance.cell(row=target_row, column=9, value=local_time)        # I Scheduled Date Local
            target_sheet_scheduled_maintenance.cell(row=target_row, column=10, value=local_tz)         # J Local Time Zone
            target_sheet_scheduled_maintenance.cell(row=target_row, column=11, value=duration)         # K Duration of change impact
            target_sheet_scheduled_maintenance.cell(row=target_row, column=12, value=window)           # L Maintenance Window
            target_sheet_scheduled_maintenance.cell(row=target_row, column=13, value=service_impact)   # M Service Impact
            target_sheet_scheduled_maintenance.cell(row=target_row, column=14, value=orange_ref)       # N Orange Change Reference
            target_sheet_scheduled_maintenance.cell(row=target_row, column=15, value=carrier_name)     # O Carrier Name
            target_sheet_scheduled_maintenance.cell(row=target_row, column=16, value=carrier_circuit)  # P Carrier Circuit ID
            target_sheet_scheduled_maintenance.cell(row=target_row, column=17, value=carrier_ref)      # Q Carrier Reference
            target_sheet_scheduled_maintenance.cell(row=target_row, column=18, value=maint_type)       # R Maintenance Type
            target_sheet_scheduled_maintenance.cell(row=target_row, column=19, value=carrier_status)   # S Carrier Status
            target_sheet_scheduled_maintenance.cell(row=target_row, column=20, value=orange_status)    # T Orange Status
            target_sheet_scheduled_maintenance.cell(row=target_row, column=24, value=carrier_link)     # U Carrier Link
            fill_row(target_sheet_scheduled_maintenance, target_row, 1, 23, CARRIER_FILL)
            # target_sheet_scheduled_maintenance.cell(row=target_row,column=21,value=ce.get_data(context, carrier_link))
            target_row += 1



def save_output():
    output_path = os.path.join(os.getcwd(), "maintainance", output_file)
    TARGET_FILE.save(output_path)
    print(f"Output saved successfully: {output_path}")



def orchestration():
    inventory_mapping()
    orange_mapping()
    carrier_mapping()
    next_row=orange_data_extraction()
    carrier_data_extraction(next_row)
    save_output()
    


orchestration()