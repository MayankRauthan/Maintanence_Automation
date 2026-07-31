from openpyxl import load_workbook
from datetime import date
import os

# ================= FILE PATHS =================
SOURCE_FILE_PATH = os.path.join(os.getcwd(), "maintainance", "Maintenance-Data(125).xlsx")
TARGET_FILE_PATH = os.path.join(os.getcwd(), "maintainance", "Haleon -Network planned maintenance tracker Week 13_25 March_2026.xlsx")
INVENTORY_FILE_PATH = os.path.join(os.getcwd(), "maintainance", "Haleon WAN Inventory.xlsx")

SOURCE_FILE = load_workbook(SOURCE_FILE_PATH, data_only=True)
TARGET_FILE = load_workbook(TARGET_FILE_PATH)
INVENTORY_FILE = load_workbook(INVENTORY_FILE_PATH, data_only=True)

# ================= SHEETS =================
target_sheet = TARGET_FILE["Scheduled Maintenance"]
inventory_sheet = INVENTORY_FILE["Haleon WAN Inventory"]

source_sheet_orange = SOURCE_FILE["Orange"]
source_sheet_carrier = SOURCE_FILE["Carrier"]

# ================= GLOBAL =================
dict_device_to_site = {}

# ================= HELPERS =================

def inventory_mapping():
    for row in inventory_sheet.iter_rows(min_row=2, values_only=True):
        if row[4] and row[0]:
            device = str(row[4]).strip().upper()
            site = str(row[0]).strip().upper()
            dict_device_to_site[device] = site


def map_service_impact(val):
    if val is None:
        return ""
    val = str(val).upper()
    if "NO" in val:
        return "No"
    return "Yes"


def write_row(sheet, row, data, sno):
    sheet.cell(row=row, column=1, value=date.today().strftime("%d-%b-%y"))
    sheet.cell(row=row, column=2, value=sno)

    sheet.cell(row=row, column=3, value=data["site"])
    sheet.cell(row=row, column=4, value=data["device"])
    sheet.cell(row=row, column=5, value=data["scope"])
    sheet.cell(row=row, column=6, value=data["type"])

    sheet.cell(row=row, column=7, value=data["city"])
    sheet.cell(row=row, column=8, value=data["country"])
    sheet.cell(row=row, column=9, value=data["region"])

    sheet.cell(row=row, column=10, value=data["sched_gmt"])
    sheet.cell(row=row, column=11, value=data["local_tz"])

    sheet.cell(row=row, column=12, value=data["duration"])
    sheet.cell(row=row, column=13, value=data["window"])

    sheet.cell(row=row, column=14, value=map_service_impact(data["impact"]))

    # Column 15–20 → Manual fields (leave blank)

    sheet.cell(row=row, column=21, value=data["status"])


# ================= ORANGE =================

def process_orange(start_row, sno):
    row_ptr = start_row

    for i in range(8, source_sheet_orange.max_row + 1):

        duration = source_sheet_orange.cell(row=i, column=19).value
        if duration and str(duration).strip().upper() == "NONE":
            continue

        device = source_sheet_orange.cell(row=i, column=6).value
        site_id = source_sheet_orange.cell(row=i, column=8).value
        scope = source_sheet_orange.cell(row=i, column=24).value
        city = source_sheet_orange.cell(row=i, column=9).value
        country = source_sheet_orange.cell(row=i, column=10).value
        region = source_sheet_orange.cell(row=i, column=11).value
        sched_gmt = source_sheet_orange.cell(row=i, column=15).value
        local_tz = source_sheet_orange.cell(row=i, column=18).value
        window = source_sheet_orange.cell(row=i, column=20).value
        impact = source_sheet_orange.cell(row=i, column=22).value
        status = source_sheet_orange.cell(row=i, column=23).value
        maint_type = source_sheet_orange.cell(row=i, column=12).value

        site = None
        if device:
            site = dict_device_to_site.get(str(device).upper())

        if not site:
            site = site_id if site_id else city

        data = {
            "site": site,
            "device": device,
            "scope": scope,
            "type": maint_type,
            "city": city,
            "country": country,
            "region": region,
            "sched_gmt": sched_gmt,
            "local_tz": local_tz,
            "duration": duration,
            "window": window,
            "impact": impact,
            "status": status
        }

        write_row(target_sheet, row_ptr, data, sno)

        row_ptr += 1
        sno += 1

    return row_ptr, sno


# ================= CARRIER =================

def process_carrier(start_row, sno):
    row_ptr = start_row

    for i in range(8, source_sheet_carrier.max_row + 1):

        device = source_sheet_carrier.cell(row=i, column=3).value
        site_id = source_sheet_carrier.cell(row=i, column=8).value
        scope = source_sheet_carrier.cell(row=i, column=30).value
        city = source_sheet_carrier.cell(row=i, column=9).value
        country = source_sheet_carrier.cell(row=i, column=10).value
        region = source_sheet_carrier.cell(row=i, column=11).value
        sched_gmt = source_sheet_carrier.cell(row=i, column=15).value
        local_tz = source_sheet_carrier.cell(row=i, column=18).value
        duration = source_sheet_carrier.cell(row=i, column=19).value
        window = source_sheet_carrier.cell(row=i, column=20).value
        impact = source_sheet_carrier.cell(row=i, column=22).value
        status = source_sheet_carrier.cell(row=i, column=29).value
        maint_type = source_sheet_carrier.cell(row=i, column=12).value

        site = None
        if device:
            site = dict_device_to_site.get(str(device).upper())

        if not site:
            site = site_id if site_id else city

        data = {
            "site": site,
            "device": device,
            "scope": scope,
            "type": maint_type,
            "city": city,
            "country": country,
            "region": region,
            "sched_gmt": sched_gmt,
            "local_tz": local_tz,
            "duration": duration,
            "window": window,
            "impact": impact,
            "status": status
        }

        write_row(target_sheet, row_ptr, data, sno)

        row_ptr += 1
        sno += 1

    return row_ptr, sno


# ================= MAIN =================

def orchestration():
    inventory_mapping()

    start_row = 12
    sno = 1

    next_row, sno = process_orange(start_row, sno)
    next_row, sno = process_carrier(next_row, sno)

    output_file = f"Haleon_Output_{date.today().isoformat()}.xlsx"
    output_path = os.path.join(os.getcwd(), "maintainance", output_file)

    TARGET_FILE.save(output_path)
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    orchestration()