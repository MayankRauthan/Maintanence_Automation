# Script Automates the maintenance report . 
# it extract data from orange and carrier sheets, maps the device to site using inventory sheet, fills the scheduled maintenance sheet in the
#  report and moves completed maintenance to completed maintenance sheet based on the date. It also updates the status of the maintenance in the report based on the data from orange and carrier sheets. Finally it saves the output with a date stamp to avoid overwriting previous outputs.
# It also preserve the formatting like color, border and row height for better readability.



# need to implement the logic when reference matches but the dates are changed as the maintenance could be reschdeduled or new status
from operator import or_

from openpyxl import load_workbook
from datetime import date, datetime, timezone , timedelta
import os
import cvmt_extraction as ce, machx_extraction as me
from playwright.sync_api import Playwright, sync_playwright
from openpyxl.styles import PatternFill, Border, Side
from copy import copy


# Define file paths
SOURCE_FILE_PATH = os.path.join(os.getcwd(), "maintainance", "Maintenance-Data (1).xlsx")
TARGET_FILE_PATH= os.path.join(os.getcwd(), "maintainance", "Haleon -Network planned maintenance tracker Week 24_08 June_2026.xlsx")
INVENTORY_FILE_PATH = os.path.join(os.getcwd(), "maintainance", "Haleon WAN Inventory.xlsx")

# Load the workbooks
SOURCE_FILE = load_workbook(SOURCE_FILE_PATH, data_only=True)
TARGET_FILE = load_workbook(TARGET_FILE_PATH, data_only=True)
INVENTORY_FILE = load_workbook(INVENTORY_FILE_PATH, data_only=True)

# Load the sheets
# Load target sheets
target_sheet_scheduled_maintenance = TARGET_FILE["Scheduled Maintenance"]
target_sheet_completed_maintenance = TARGET_FILE["Completed Maintenance"]

# Load inventory sheet
inventory_sheet = INVENTORY_FILE["Haleon WAN Inventory"]    


#Load source sheets
source_sheet_orange = SOURCE_FILE["Orange"]
source_sheet_carrier = SOURCE_FILE["Carrier"]
cover_sheet = TARGET_FILE["Cover"]

# save the file name with date added ( YYYY-MM-DD ) to avoid overwriting previous outputs
output_file = f"Haleon_Automated_Maintenance_Output_{date.today().isoformat()}.xlsx"

#initialize the mapping dictionary for device to site mapping
dict_device_to_site = {}

orange_mapping_dict = {}

carrier_mapping_dict = {}

or_dict ={}
cr_dict= {}
report_dict={}

rows_to_delete = set()



CARRIER_FILL = PatternFill(fill_type="solid", fgColor="DCE6F1")   # light blue
ORANGE_FILL = PatternFill(fill_type="solid", fgColor="F2DCDB")  # light pink
Blue_Fill = PatternFill(fill_type="solid", fgColor="538DD5")    # light blue
white_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")    # white
Yellow_Fill = PatternFill(fill_type="solid", fgColor="FFFF00")    # yellow
Red_fill = PatternFill(fill_type="solid", fgColor="FF0000")    # red

thin = Side(style="thin", color="000000")
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)


# maps device name to site name using the inventory sheet and stores it in a dictionary for quick lookup 
def inventory_mapping():
    for row in inventory_sheet.iter_rows(min_row=2, values_only=True):
        device_name = row[4].strip().upper()
        site_name = row[0].strip().upper()
        #print ( f"mapping device {device_name} to site {site_name}")
        dict_device_to_site[device_name] = site_name



def orange_mapping():
    global orange_mapping_dict
    orange_mapping_dict = {
    "orange_ref": 1,          # A
    "device_name": 3,        
    "indirect device_name": 6,   # D
    "site_id": 8,             # H
    "city": 9,                # I
    "country": 10,           # J
    "region": 11,            # K
    "maint_type": 12,        # L
    "sched_gmt": 15,         # O
    "local_time": 17,        # Q
    "local_tz": 18,          # R
    "duration": 19,          # S
    "window": 20,            # T
    "service_impact": 22,    # V
    "orange_status": 23,     # W
    "carrier_status": 29,
    "scope": 24,             # X
}

def carrier_mapping():
   global carrier_mapping_dict
   carrier_mapping_dict = {
    "orange_ref": 1,          # A (row[0])
    "device_name": 3,        # C (row[2])
    "indirect device_name": 6,
    "site_id": 8,             # H (row[7])
    "city": 9,                # I (row[8])
    "country": 10,           # J (row[9])
    "region": 11,            # K (row[10])
    "maint_type": 12,        # L (row[11])
    "sched_gmt": 15,         # O (row[14])
    "local_time": 17,        # Q (row[16])
    "local_tz": 18,          # R (row[17])
    "duration": 19,          # S (row[18])
    "window": 20,            # T (row[19])
    "service_impact": 22,    # V (row[21])
    "orange_status": 23,     # W (row[22])    23
    "carrier_name": 26,     # Z (row[25])
    "carrier_circuit": 27,  # AA (row[26])
    "carrier_ref": 28,      # AB (row[27])
    "carrier_status": 29,   # AC (row[28])    19
    "reason": 30,            # AD (row[29])
}



def extract_status():
    with sync_playwright() as p:
        browser = p.firefox.launch(headless=False)
        context = browser.new_context()
        for i in range (8, source_sheet_orange.max_row + 1):
            cell_value = source_sheet_orange.cell(row=i, column=34).value
            if cell_value:
                print (f"Status found in Orange sheet at row {i}: {cell_value}")
                print(me.get_data(context, cell_value))
            

def fill_row(sheet, row, start_col, end_col, fill):
    for col in range(start_col, end_col + 1):
        sheet.cell(row=row, column=col).fill = fill
        sheet.cell(row=row, column=col).border =  thin_border


def change_status():
    global  or_dict 
    global  cr_dict
    global report_dict
  
    for i in range(8, source_sheet_orange.max_row + 1):  # create a hashmap of orange_ref to its date and row number for quick lookup
        orange_ref = source_sheet_orange.cell(row=i, column=orange_mapping_dict["orange_ref"]).value
        orange_device = source_sheet_orange.cell(row=i, column=orange_mapping_dict["device_name"]).value
        print(f"processing orange ref {orange_ref} with device name {orange_device} for change status")
        if orange_device is None or orange_device.strip() == "": # if device name is not present in orange sheet then take the indirect device name for matching with report
            orange_device=source_sheet_orange.cell(row=i, column=orange_mapping_dict["indirect device_name"]).value
            print(f"device name is None for orange ref {orange_ref} taking indirect device name {orange_device} for matching with report")
        orange_date= source_sheet_orange.cell(row=i, column=orange_mapping_dict["sched_gmt"]).value
        or_dict[orange_ref + orange_device]=[orange_date,i]

    for j in range(8, source_sheet_carrier.max_row + 1): # create a hashmap of carrier_ref to its date and row number for quick lookup
        carrier_ref = source_sheet_carrier.cell(row=j, column=carrier_mapping_dict["orange_ref"]).value
        carrier_date= source_sheet_carrier.cell(row=j, column=carrier_mapping_dict["sched_gmt"]).value
        carrier_device= source_sheet_carrier.cell(row=j, column=carrier_mapping_dict["device_name"]).value # if device name is not present in carrier sheet then take the indirect device name for matching with report
        if carrier_device is None or carrier_device.strip() == "":
            carrier_device=source_sheet_carrier.cell(row=j, column=carrier_mapping_dict["indirect device_name"]).value
        cr_dict[carrier_ref + carrier_device]=[carrier_date,j]

    for k in range(12, target_sheet_scheduled_maintenance.max_row + 1): # create a hashmap of report_ref to its date and row number for quick lookup
        report_ref = target_sheet_scheduled_maintenance.cell(row=k, column=14).value
        report_date= target_sheet_scheduled_maintenance.cell(row=k, column=7).value
        report_device= target_sheet_scheduled_maintenance.cell(row=k, column=2).value
        report_dict[report_ref + report_device]=[report_date,k]
        print(f"city value is {target_sheet_scheduled_maintenance.cell(row=k, column=4).value} and country value is {target_sheet_scheduled_maintenance.cell(row=k, column=5).value} for report ref {report_ref} in report hashmap")
   
    # for ref in report_dict.keys(): # compare the dates and update the status in the report accordingly
    #     #print(f"Checking report ref {ref} with date {report_dict[ref][0]} and row number {report_dict[ref][1]} in report hashmap against orange and carrier hashmaps")
    #     report_date=report_dict[ref][0]
    #     report_row=report_dict[ref][1]
    #     print(cr_dict.keys())
    #     print("-----------------")
    #     print(or_dict.keys())
    #     if ref in or_dict.keys():
    #         orange_date=or_dict[ref][0]
    #         print(f"removed orange ref {ref} with date {orange_date} and row number {or_dict[ref][1]} from hashmap")

    #         or_dict.pop(ref) # remove the matched ref from hashmap to optimize further lookups

    #         if orange_date!=report_date:
    #             target_sheet_scheduled_maintenance.cell(row=report_row, column=21, value="Rescheduled")
    #         else:
    #             target_sheet_scheduled_maintenance.cell(row=report_row, column=21, value="Already Notified")
    #             target_sheet_scheduled_maintenance.cell(row=report_row, column=21).fill = white_fill
    #         target_sheet_completed_maintenance.cell(row=report_row, column=20).value= source_sheet_orange.cell(row=or_dict[ref][1], column=orange_mapping_dict["orange_status"]).value

    #     elif ref in cr_dict.keys():
    #         carrier_date=cr_dict[ref][0]
    #         val =cr_dict.pop(ref) # remove the matched ref from hashmap to optimize further lookups
    #         print(f"removed carrier ref {ref} with date {val[0]} and row number {val[1]} from hashmap")
    #         if carrier_date!=report_date:
    #             target_sheet_scheduled_maintenance.cell(row=report_row, column=21, value="Rescheduled")
                
    #         else:
    #             target_sheet_scheduled_maintenance.cell(row=report_row, column=21, value="Already Notified")
    #             target_sheet_scheduled_maintenance.cell(row=report_row, column=21).fill = white_fill
    #         target_sheet_completed_maintenance.cell(row=report_row, column=20).value= source_sheet_carrier.cell(row=val[1], column=carrier_mapping_dict["orange_status"]).value


def move_completed_maintenance():
    print("Running move_completed_maintenance function to move completed maintenance from Scheduled Maintenance sheet to Completed Maintenance sheet based on the date")
    current_row = 12
    max_row = target_sheet_scheduled_maintenance.max_row
    target_row = target_sheet_completed_maintenance.max_row + 1

    while current_row <= max_row:
        date=target_sheet_scheduled_maintenance.cell(row=current_row, column=7).value
        date= datetime.strptime(date, "%d/%b/%Y %I:%M:%S %p")
        date = date.replace(tzinfo=timezone.utc)
        now_time = datetime.now(timezone.utc) - timedelta(days=1)
        now_time = now_time.replace(hour=23, minute=59, second=59)   

       # print(f"Checking maintenance at row {current_row} with date {date} and today's date {now_time}")
        if date <= now_time:
            #print("   its smaller moving ")
            for col in range(1, 24):
                value = target_sheet_scheduled_maintenance.cell(row=current_row, column=col).value
                target_sheet_completed_maintenance.cell(row=target_row, column=col, value=value)
            fill_row(target_sheet_completed_maintenance, target_row, 1, 23, copy(target_sheet_scheduled_maintenance.cell(row=current_row, column=1).fill))
            target_sheet_completed_maintenance.row_dimensions[target_row].height = 30  # set row height for better readability
            target_row += 1
            target_sheet_scheduled_maintenance.delete_rows(current_row)
           # print (f"Moved completed maintenance with reference {target_sheet_scheduled_maintenance.cell(row=current_row, column=14).value} to Completed Maintenance sheet")
            max_row -= 1
        else:            
            current_row += 1
            


def orange_data_extraction():
    target_row = target_sheet_scheduled_maintenance.max_row + 1
    
    for values in or_dict.values():
        print (f"Remaining Orange Ref in hashmap with date {values[0]} and row number {values[1]}")
        i=values[1]
        duration_value = source_sheet_orange.cell(
            row=i,
            column=orange_mapping_dict["duration"]
        ).value

        # Skip rows where duration is literal text "NONE"
        if duration_value is not None and str(duration_value).strip().upper() == "NONE":
            continue
        
        device_name = source_sheet_orange.cell(row=i, column=orange_mapping_dict["device_name"]).value
        if device_name is None or device_name.strip() == "": # if device name is not present in orange sheet then take the indirect device name for matching with report
            device_name=source_sheet_orange.cell(row=i, column=orange_mapping_dict["indirect device_name"]).value   
        site_id = source_sheet_orange.cell(row=i, column=orange_mapping_dict["site_id"]).value
        scope = source_sheet_orange.cell(row=i, column=orange_mapping_dict["scope"]).value
        city = source_sheet_orange.cell(row=i, column=orange_mapping_dict["city"]).value
        country = source_sheet_orange.cell(row=i, column=orange_mapping_dict["country"]).value
        region = source_sheet_orange.cell(row=i, column=orange_mapping_dict["region"]).value
        sched_gmt = source_sheet_orange.cell(row=i, column=orange_mapping_dict["sched_gmt"]).value
        local_time = source_sheet_orange.cell(row=i, column=orange_mapping_dict["local_time"]).value
        local_tz = source_sheet_orange.cell(row=i, column=orange_mapping_dict["local_tz"]).value
        duration = source_sheet_orange.cell(row=i, column=orange_mapping_dict["duration"]).value
        window = source_sheet_orange.cell(row=i, column=orange_mapping_dict["window"]).value
        service_impact = source_sheet_orange.cell(row=i, column=orange_mapping_dict["service_impact"]).value
        orange_ref = source_sheet_orange.cell(row=i, column=orange_mapping_dict["orange_ref"]).value
        maint_type = source_sheet_orange.cell(row=i, column=orange_mapping_dict["maint_type"]).value
        orange_status = source_sheet_orange.cell(row=i, column=orange_mapping_dict["orange_status"]).value
        orange_link= source_sheet_orange.cell(row=i, column=34).value
        carrier_status = source_sheet_orange.cell(row=i, column=orange_mapping_dict["carrier_status"]).value
        # Site name from inventory, fallback to site_id from orange sheet
        site_name = None
        if device_name is not None:
            site_name = dict_device_to_site.get(str(device_name).strip().upper())

        if site_name is None:
            site_name=city
        # if not site_name and site_id is not None:
        #     site_name = site_id
        # elif site_name is None:
        #     site_name=city

        #print(f"country value is {country} and region value is {region} for orange ref {orange_ref}")

        # Write to Scheduled Maintenance sheet
        target_sheet_scheduled_maintenance.cell(row=target_row, column=1, value=site_name)        # A Site Name
        target_sheet_scheduled_maintenance.cell(row=target_row, column=2, value=device_name)      # B Device Name
        target_sheet_scheduled_maintenance.cell(row=target_row, column=3, value=scope)            # C Scope of Change
        target_sheet_scheduled_maintenance.cell(row=target_row, column=4, value=city)             # D City
        target_sheet_scheduled_maintenance.cell(row=target_row, column=5, value=country)          # E County/Country
        target_sheet_scheduled_maintenance.cell(row=target_row, column=6, value=region)           # F Region
        target_sheet_scheduled_maintenance.cell(row=target_row, column=7, value=sched_gmt)        # G Scheduled Date GMT
        target_sheet_scheduled_maintenance.cell(row=target_row, column=9, value=local_time)       # I Scheduled Date Local
        target_sheet_scheduled_maintenance.cell(row=target_row, column=10, value=local_tz)        # J Local Time Zone
        target_sheet_scheduled_maintenance.cell(row=target_row, column=11, value=duration)        # K Duration
        target_sheet_scheduled_maintenance.cell(row=target_row, column=12, value=window)          # L Maintenance Window
        target_sheet_scheduled_maintenance.cell(row=target_row, column=13, value=service_impact)  # M Service Impact
        target_sheet_scheduled_maintenance.cell(row=target_row, column=14, value=orange_ref)      # N Orange Change Reference
        target_sheet_scheduled_maintenance.cell(row=target_row, column=18, value=maint_type)      # R Maintenance Type
        target_sheet_scheduled_maintenance.cell(row=target_row, column=20, value=orange_status)   # T Orange Status
        target_sheet_scheduled_maintenance.cell(row=target_row, column=24, value=orange_link) 
        target_sheet_scheduled_maintenance.cell(row=target_row, column=19, value=carrier_status)   # U Carrier Status
        fill_row(target_sheet_scheduled_maintenance, target_row, 1, 23, ORANGE_FILL)
        target_sheet_scheduled_maintenance.row_dimensions[target_row].height = 30  # set row height for better readability  
       
        # Logic for maintenance status 
        ref=orange_ref + device_name
        maint_type=maint_type.strip()
        if ref in report_dict.keys():
            rows_to_delete.add(report_dict[ref][1]) # add the matched ref row number from orange sheet to a list for deletion later to optimize the hashmap lookups
            sched_gmt= datetime.strptime(sched_gmt, "%d/%b/%Y %I:%M:%S %p")
            sched_gmt = sched_gmt.replace(tzinfo=timezone.utc)
            target_date=report_dict[ref][0]
            target_date=datetime.strptime(target_date, "%d/%b/%Y %I:%M:%S %p")
            target_date = target_date.replace(tzinfo=timezone.utc)

            #copy last two columnd from report which gurmeet maually enters
            target_sheet_scheduled_maintenance.cell(row=target_row, column=22, value=target_sheet_scheduled_maintenance.cell(row=report_dict[ref][1], column=22).value) # V Responsible Team
            target_sheet_scheduled_maintenance.cell(row=target_row, column=23, value=target_sheet_scheduled_maintenance.cell(row=report_dict[ref][1], column=23).value) # W Maintenance Description

            if(target_date==sched_gmt):
                target_sheet_scheduled_maintenance.cell(row=target_row, column=21, value="Already Notified")
                target_sheet_scheduled_maintenance.cell(row=target_row, column=21).fill = white_fill
            else:
                target_sheet_scheduled_maintenance.cell(row=target_row, column=21, value="Rescheduled")
                target_sheet_scheduled_maintenance.cell(row=target_row, column=21).fill = white_fill
        elif str.upper(maint_type) =="EXPEDITE MAINTENANCE":
            target_sheet_scheduled_maintenance.cell(row=target_row, column=21, value="EXPEDITE MAINTENANCE")
            target_sheet_scheduled_maintenance.cell(row=target_row, column=21).fill = Yellow_Fill
        
        elif str.upper(maint_type) =="EMERGENCY MAINTENANCE":
            target_sheet_scheduled_maintenance.cell(row=target_row, column=21, value="CANCELLED")
            target_sheet_scheduled_maintenance.cell(row=target_row, column=21).fill = Red_fill
        else:
            target_sheet_scheduled_maintenance.cell(row=target_row, column=21, value="New Maintenance")
            target_sheet_scheduled_maintenance.cell(row=target_row, column=21).fill = Blue_Fill

        if str.upper(carrier_status)=="CANCELLED" or str.upper(orange_status)=="CANCELED":
            target_sheet_scheduled_maintenance.cell(row=target_row, column=21, value="CANCELLED")
            target_sheet_scheduled_maintenance.cell(row=target_row, column=21).fill = Red_fill
        

        
        target_row += 1
        
    return target_row

def carrier_data_extraction(start_row):
    target_row = start_row

    
    for values in cr_dict.values():
        i=values[1]
        print (f"Remaining Carrier Ref in hashmap with date {values[0]} and row number {values[1]}")
        device_name = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["device_name"]).value
        if device_name is None or device_name.strip() == "": # if device name is not present in carrier sheet then take the indirect device name for matching with report
            device_name=source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["indirect device_name"]).value
        site_id = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["site_id"]).value
        scope = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["reason"]).value
        city = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["city"]).value
        country = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["country"]).value
        region = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["region"]).value
        sched_gmt = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["sched_gmt"]).value
        local_time = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["local_time"]).value
        local_tz = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["local_tz"]).value
        duration = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["duration"]).value
        window = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["window"]).value
        service_impact = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["service_impact"]).value
        orange_ref = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["orange_ref"]).value
        carrier_name = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["carrier_name"]).value
        carrier_circuit = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["carrier_circuit"]).value
        carrier_ref = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["carrier_ref"]).value
        carrier_status = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["carrier_status"]).value
        maint_type = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["maint_type"]).value
        orange_status = source_sheet_carrier.cell(row=i, column=carrier_mapping_dict["orange_status"]).value
        carrier_link = source_sheet_carrier.cell(row=i, column=34).value

        # Site name from inventory, fallback to site_id from carrier sheet
        site_name = None
        if device_name is not None:
            site_name = dict_device_to_site.get(str(device_name).strip().upper())
           # print(f"site name from inventory for device {device_name} is {site_name} and site id from carrier sheet is {site_id}")
            #print(f"not site name {not site_name}")
        # if not site_name and site_id is not None:
        #     site_name = site_id
        # elif site_name is None: 
        #     site_name=city

        if site_name is None:
            site_name=city
        
        #print(site_name)
        #print(f"country value is {country} and region value is {region} for orange ref {orange_ref}")

        # Write to Scheduled Maintenance sheet
        target_sheet_scheduled_maintenance.cell(row=target_row, column=1, value=site_name)         # A Site Name
        target_sheet_scheduled_maintenance.cell(row=target_row, column=2, value=device_name)       # B Device Name
        target_sheet_scheduled_maintenance.cell(row=target_row, column=3, value=scope)             # C Scope of Change
        target_sheet_scheduled_maintenance.cell(row=target_row, column=4, value=city)              # D City
        target_sheet_scheduled_maintenance.cell(row=target_row, column=5, value=country)           # E County/Country
        target_sheet_scheduled_maintenance.cell(row=target_row, column=6, value=region)            # F Region
        target_sheet_scheduled_maintenance.cell(row=target_row, column=7, value=sched_gmt)         # G Scheduled Date GMT
        target_sheet_scheduled_maintenance.cell(row=target_row, column=9, value=local_time)        # I Scheduled Date Local
        target_sheet_scheduled_maintenance.cell(row=target_row, column=10, value=local_tz)         # J Local Time Zone
        target_sheet_scheduled_maintenance.cell(row=target_row, column=11, value=duration)         # K Duration of change impact
        target_sheet_scheduled_maintenance.cell(row=target_row, column=12, value=window)           # L Maintenance Window
        target_sheet_scheduled_maintenance.cell(row=target_row, column=13, value=service_impact)   # M Service Impact
        target_sheet_scheduled_maintenance.cell(row=target_row, column=14, value=orange_ref)       # N Orange Change Reference
        target_sheet_scheduled_maintenance.cell(row=target_row, column=15, value=carrier_name)     # O Carrier Name
        target_sheet_scheduled_maintenance.cell(row=target_row, column=16, value=carrier_circuit)  # P Carrier Circuit ID
        target_sheet_scheduled_maintenance.cell(row=target_row, column=17, value=carrier_ref)      # Q Carrier Reference
        target_sheet_scheduled_maintenance.cell(row=target_row, column=18, value=maint_type)       # R Maintenance Type
        target_sheet_scheduled_maintenance.cell(row=target_row, column=19, value=carrier_status)   # S Carrier Status
        target_sheet_scheduled_maintenance.cell(row=target_row, column=20, value=orange_status)    # T Orange Status
        target_sheet_scheduled_maintenance.cell(row=target_row, column=24, value=carrier_link) 
        target_sheet_scheduled_maintenance.cell(row=target_row, column=21).value = "New Maintenance" 
        target_sheet_scheduled_maintenance.cell(row=target_row, column=21).fill = Blue_Fill  # U Carrier Link
           # U Carrier Link
        fill_row(target_sheet_scheduled_maintenance, target_row, 1, 23, CARRIER_FILL)
      
        target_sheet_scheduled_maintenance.row_dimensions[target_row].height = 30  # set row height for better readability  
        
        ref=orange_ref + device_name
        maint_type=maint_type.strip()
        if ref in report_dict.keys():
            rows_to_delete.add(report_dict[ref][1]) # add the matched ref row number from orange sheet to a set for deletion later to optimize the hashmap lookups
            sched_gmt= datetime.strptime(sched_gmt, "%d/%b/%Y %I:%M:%S %p")
            sched_gmt = sched_gmt.replace(tzinfo=timezone.utc)
            target_date=report_dict[ref][0]
            target_date=datetime.strptime(target_date, "%d/%b/%Y %I:%M:%S %p")
            target_date = target_date.replace(tzinfo=timezone.utc)

            # copy last two columnd from report which gurmeet maually enters
            target_sheet_scheduled_maintenance.cell(row=target_row, column=22, value=target_sheet_scheduled_maintenance.cell(row=report_dict[ref][1], column=22).value) # V Responsible Team
            target_sheet_scheduled_maintenance.cell(row=target_row, column=23, value=target_sheet_scheduled_maintenance.cell(row=report_dict[ref][1], column=23).value) # W Maintenance Description

            if(target_date==sched_gmt):
                target_sheet_scheduled_maintenance.cell(row=target_row, column=21, value="Already Notified")
                target_sheet_scheduled_maintenance.cell(row=target_row, column=21).fill = white_fill
            else:
                target_sheet_scheduled_maintenance.cell(row=target_row, column=21, value="Rescheduled")
                target_sheet_scheduled_maintenance.cell(row=target_row, column=21).fill = white_fill
        elif str.upper(maint_type) =="EXPEDITE MAINTENANCE":
            target_sheet_scheduled_maintenance.cell(row=target_row, column=21, value="EXPEDITE MAINTENANCE")
            target_sheet_scheduled_maintenance.cell(row=target_row, column=21).fill = Yellow_Fill
        elif str.upper(maint_type) =="EMERGENCY MAINTENANCE":
            target_sheet_scheduled_maintenance.cell(row=target_row, column=21, value="CANCELLED")
            target_sheet_scheduled_maintenance.cell(row=target_row, column=21).fill = Red_fill
        else:
            target_sheet_scheduled_maintenance.cell(row=target_row, column=21, value="New Maintenance")
            target_sheet_scheduled_maintenance.cell(row=target_row, column=21).fill = Blue_Fill

        if str.upper(carrier_status)=="CANCEL/CANCELLED" or str.upper(orange_status)=="CANCELED":
            target_sheet_scheduled_maintenance.cell(row=target_row, column=21, value="CANCELLED")
            target_sheet_scheduled_maintenance.cell(row=target_row, column=21).fill = Yellow_Fill

        target_row += 1

def delete_rows():
    for row in sorted(rows_to_delete, reverse=True):
        print(f"Deleting row {row} from Scheduled Maintenance sheet as the maintenance with reference {target_sheet_scheduled_maintenance.cell(row=row, column=14).value} is either rescheduled or cancelled or already notified to avoid duplicates in the report")
        target_sheet_scheduled_maintenance.delete_rows(row)

def save_output():
    output_path = os.path.join(os.getcwd(), "maintainance", output_file)
    TARGET_FILE.save(output_path)
    print(f"Output saved successfully: {output_path}")

def update_no():
    already_notified_count = 0
    new_maintenance_count = 0
    rescheduled_count = 0
    expedite_maintenance_count = 0
    cancelled_count = 0
    for row in range(12, target_sheet_scheduled_maintenance.max_row + 1):
        if target_sheet_scheduled_maintenance.cell(row=row, column=21).value == "Already Notified":
            already_notified_count += 1
        if target_sheet_scheduled_maintenance.cell(row=row, column=21).value == "New Maintenance":
            new_maintenance_count += 1
        if target_sheet_scheduled_maintenance.cell(row=row, column=21).value =="Rescheduled":
           rescheduled_count += 1
        if target_sheet_scheduled_maintenance.cell(row=row, column=21).value =="EXPEDITE MAINTENANCE":
           expedite_maintenance_count += 1
        if target_sheet_scheduled_maintenance.cell(row=row, column=21).value =="CANCELLED":
           cancelled_count += 1 
    target_sheet_scheduled_maintenance.cell(row=4, column=5, value=already_notified_count) # B2 Already Notified Count
    target_sheet_scheduled_maintenance.cell(row=5, column=5, value=rescheduled_count) # C2 New Maintenance Count
    target_sheet_scheduled_maintenance.cell(row=3, column=5, value=new_maintenance_count) # D2 Rescheduled Count
    target_sheet_scheduled_maintenance.cell(row=6, column=5, value=cancelled_count) # D2 Rescheduled Count
    target_sheet_scheduled_maintenance.cell(row=7, column=5, value=expedite_maintenance_count) # E2 Expedite Maintenance Count

def update_cover_date_week():
    if "Cover" not in TARGET_FILE.sheetnames:
        return
    now = datetime.now()
    week_no = now.isocalendar()[1]
    cover_sheet["B9"] = f"{now.strftime('%d-%b-%Y')} - Week {week_no}"

def orchestration():
    inventory_mapping()
    orange_mapping()
    carrier_mapping()
    change_status() 
    next_row=orange_data_extraction()
    carrier_data_extraction(next_row)
    delete_rows()
    move_completed_maintenance()
    update_no()
    update_cover_date_week()
    save_output()
    


orchestration()